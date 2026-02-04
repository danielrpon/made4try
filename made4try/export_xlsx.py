# Exportar a XLSX y embeber HTML
# export_xlsx.py
from io import BytesIO
import pandas as pd
from .config import DEFAULT_SHEET_NAME


def dataframe_to_xlsx_bytes(
    df: pd.DataFrame,
    html_chart: str = None,
    sheet_name: str = DEFAULT_SHEET_NAME
) -> BytesIO:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        # ------------------- Hoja principal (datos) -------------------
        df.to_excel(xw, index=False, sheet_name=sheet_name)
        ws = xw.book[sheet_name]

        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment

        # Anchos base + métricas clásicas
        widths = {
            "fecha": 18, "documento": 24, "elapsed_s": 12, "power_w": 12, "hr_bpm": 12,
            "speed_kmh": 12, "dt_s": 12,
        }

        metrics = {
            "pct_ftp", "pct_fc_rel", "EFR", "IF", "ICR",
            "TSS_inc", "FSS_inc", "TSS_inc_ma30", "FSS_inc_ma30",
            "power_ma30", "hr_ma30", "TSS", "FSS", "TSS_total", "FSS_total",
        }

        # Nuevas columnas (ventana EF/DA)
        window_cols = {
            "WIN_mode", "WIN_mins", "WIN_signal", "WIN_start_s", "WIN_end_s",
            "WIN_score", "WIN_cv_intensity", "WIN_hr_cov", "WIN_reason",
            "EF_win", "DA_win_pct", "EF_half1", "EF_half2",
        }

        for i, col in enumerate(df.columns, start=1):
            if col in widths:
                w = widths[col]
            elif col in window_cols:
                # un poquito más ancho para textos
                w = 16 if col.startswith("WIN_") else 14
            elif col in metrics:
                w = 14
            else:
                w = 12
            ws.column_dimensions[get_column_letter(i)].width = w

        # Header en negrita
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # ------------------- Hoja Resumen -------------------
        summary = xw.book.create_sheet("Resumen")
        summary["A1"] = "Resumen del Entrenamiento"
        summary["A1"].font = Font(bold=True, size=14)

        def _get_first(col, default=None):
            try:
                if col in df.columns and len(df):
                    v = df[col].iloc[0]
                    return v if str(v) != "nan" else default
            except Exception:
                pass
            return default

        def _get_float(col, default=None):
            v = _get_first(col, default)
            try:
                return float(v)
            except Exception:
                return default

        # Datos base
        fecha = _get_first("fecha")
        doc = _get_first("documento")
        tss_total = _get_float("TSS_total")
        fss_total = _get_float("FSS_total")

        # Ventana EF/DA
        win_mode = _get_first("WIN_mode")
        win_mins = _get_first("WIN_mins")
        win_signal = _get_first("WIN_signal")
        win_start = _get_float("WIN_start_s")
        win_end = _get_float("WIN_end_s")
        win_score = _get_float("WIN_score")
        win_cv = _get_float("WIN_cv_intensity")
        win_hr_cov = _get_float("WIN_hr_cov")
        win_reason = _get_first("WIN_reason")

        ef_win = _get_float("EF_win")
        da_win = _get_float("DA_win_pct")
        ef1 = _get_float("EF_half1")
        ef2 = _get_float("EF_half2")

        rows = [
            ("Archivo", doc),
            ("Fecha", fecha),
            ("TSS Total", f"{tss_total:.1f}" if tss_total is not None else None),
            ("FSS Total", f"{fss_total:.1f}" if fss_total is not None else None),
            ("", ""),
            ("Ventana (modo)", win_mode),
            ("Ventana objetivo (min)", f"{float(win_mins):.0f}" if win_mins is not None and str(win_mins) != "nan" else None),
            ("Señal selección", win_signal),
            ("Inicio ventana (s)", f"{win_start:.0f}" if win_start is not None else None),
            ("Fin ventana (s)", f"{win_end:.0f}" if win_end is not None else None),
            ("Score (ventana)", f"{win_score:.4f}" if win_score is not None else None),
            ("CV intensidad", f"{win_cv:.4f}" if win_cv is not None else None),
            ("HR coverage", f"{win_hr_cov:.2f}" if win_hr_cov is not None else None),
            ("Motivo si no aplica", win_reason),
            ("", ""),
            ("EF (ventana)", f"{ef_win:.5f}" if ef_win is not None else None),
            ("DA % (ventana)", f"{da_win:.2f}%" if da_win is not None else None),
            ("EF mitad 1", f"{ef1:.5f}" if ef1 is not None else None),
            ("EF mitad 2", f"{ef2:.5f}" if ef2 is not None else None),
        ]

        start_row = 3
        for r, (k, v) in enumerate(rows, start=start_row):
            summary[f"A{r}"] = k
            summary[f"B{r}"] = v

        summary.column_dimensions["A"].width = 26
        summary.column_dimensions["B"].width = 40

        for r in range(start_row, start_row + len(rows)):
            summary[f"A{r}"].font = Font(bold=True)
            summary[f"A{r}"].alignment = Alignment(horizontal="left")
            summary[f"B{r}"].alignment = Alignment(horizontal="left")

        # ------------------- Hoja Gráficas (HTML embebido) -------------------
        if html_chart:
            chart_sheet = xw.book.create_sheet("Gráficas")
            chart_sheet["A1"] = "Gráfica Interactiva de Carga"
            chart_sheet["A1"].font = Font(bold=True, size=12)
            chart_sheet["A3"] = "Descarga el archivo HTML adjunto para ver la gráfica."
            preview = (html_chart[:500] + "...") if len(html_chart) > 500 else html_chart
            chart_sheet["A5"] = "Vista previa (HTML):"
            chart_sheet["A6"] = preview
            chart_sheet.column_dimensions["A"].width = 100

    bio.seek(0)
    return bio
