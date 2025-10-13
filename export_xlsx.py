# Exportar a XLSX y embeber HTML
# export_xlsx.py
from io import BytesIO
import pandas as pd
from .config import DEFAULT_SHEET_NAME

def dataframe_to_xlsx_bytes(df: pd.DataFrame, html_chart: str = None, sheet_name: str = DEFAULT_SHEET_NAME) -> BytesIO:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name=sheet_name)
        ws = xw.book[sheet_name]

        from openpyxl.utils import get_column_letter
        widths = {
            "fecha":18,"documento":20,"elapsed_s":12,"power_w":12,"hr_bpm":12,"speed_kmh":12,"dt_s":12,
        }
        metrics = {"pct_ftp","pct_fc_rel","EFR","IF","ICR","TSS_inc","FSS_inc","TSS_inc_ma30","FSS_inc_ma30",
                   "power_ma30","hr_ma30","TSS","FSS","TSS_total","FSS_total"}
        for i,col in enumerate(df.columns, start=1):
            w = widths.get(col, 14 if col in metrics else 12)
            ws.column_dimensions[get_column_letter(i)].width = w

        if html_chart:
            chart_sheet = xw.book.create_sheet("Gráficas")
            chart_sheet["A1"] = "Gráfica Interactiva de Carga"
            chart_sheet["A3"] = "Descarga el archivo HTML adjunto para ver la gráfica."
            preview = (html_chart[:500] + "...") if len(html_chart) > 500 else html_chart
            chart_sheet["A5"] = "Vista previa (HTML):"
            chart_sheet["A6"] = preview
            chart_sheet.column_dimensions['A'].width = 100
    bio.seek(0)
    return bio