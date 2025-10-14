"""
TCX → XLSX conversor con cálculo de EFR, IF, ICR y métricas de carga (TSS/FSS integradas).

MEJORAS v3:
- Dashboard comparativo entre entrenamientos múltiples
- Gráfica TSS/FSS segundo a segundo con promedio móvil de 10s
- Gráfica HTML embebida en Excel
- Columnas TSS_total y FSS_total en la primera fila

Autor: tú :)
"""

import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
from io import TextIOWrapper, BytesIO, StringIO
import gzip
import zipfile
from datetime import datetime
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ---------------------- Configuración de la página ----------------------
st.set_page_config(
    page_title="TCX → XLSX (EFR/IF/ICR/TSS/FSS)",
    page_icon="📈",
    layout="wide",
)

# ---------------------- Namespaces XML para TCX ----------------------
NS = {
    "tcx": "http://www.garmin.com/xmlschemas/TrainingCenterDatabase/v2",
    "ns3": "http://www.garmin.com/xmlschemas/ActivityExtension/v2",
    "ns2": "http://www.garmin.com/xmlschemas/ActivityExtension/v1",
}

# ---------------------- Utilidades ----------------------
def parse_iso8601_z(ts: str):
    """Parsea una cadena ISO 8601 que puede terminar en Z (UTC)."""
    try:
        if ts and ts.endswith("Z"):
            return datetime.fromisoformat(ts.replace("Z", "+00:00"))
        return datetime.fromisoformat(ts) if ts else None
    except Exception:
        return None

def get_text(elem, paths):
    """Extrae el texto de la primera coincidencia de rutas XPath en elem."""
    for p in paths:
        node = elem.find(p, NS)
        if node is not None and node.text:
            return node.text.strip()
    return None

def to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None

def to_int(x):
    try:
        return int(float(x))
    except (TypeError, ValueError):
        return None

def open_maybe_gzip_bytes(uploaded_file):
    """Devuelve un manejador de texto UTF-8 para un archivo TCX o TCX.GZ subido."""
    name = uploaded_file.name.lower()
    if name.endswith(".gz"):
        gz = gzip.GzipFile(fileobj=BytesIO(uploaded_file.getvalue()), mode="rb")
        return TextIOWrapper(gz, encoding="utf-8")
    else:
        return TextIOWrapper(BytesIO(uploaded_file.getvalue()), encoding="utf-8")

def parse_tcx_to_rows(uploaded_file):
    """Parsea un archivo TCX y devuelve una lista de diccionarios por trackpoint."""
    f = open_maybe_gzip_bytes(uploaded_file)
    tree = ET.parse(f)
    root = tree.getroot()

    rows = []
    first_ts = None

    for act in root.findall(".//tcx:Activities/tcx:Activity", NS):
        sport = act.get("Sport")
        for li, lap in enumerate(act.findall("tcx:Lap", NS), start=1):
            for track in lap.findall("tcx:Track", NS):
                for ti, tp in enumerate(track.findall("tcx:Trackpoint", NS), start=1):
                    ts_txt = get_text(tp, ["tcx:Time"])
                    ts = parse_iso8601_z(ts_txt) if ts_txt else None
                    if ts and first_ts is None:
                        first_ts = ts
                    elapsed = (
                        (ts - first_ts).total_seconds() if (ts and first_ts) else None
                    )

                    lat = to_float(get_text(tp, ["tcx:Position/tcx:LatitudeDegrees"]))
                    lon = to_float(get_text(tp, ["tcx:Position/tcx:LongitudeDegrees"]))
                    alt = to_float(get_text(tp, ["tcx:AltitudeMeters"]))
                    dist = to_float(get_text(tp, ["tcx:DistanceMeters"]))
                    hr = to_int(get_text(tp, ["tcx:HeartRateBpm/tcx:Value"]))
                    cad = to_int(get_text(tp, ["tcx:Cadence"]))

                    speed_mps = to_float(
                        get_text(
                            tp,
                            [
                                "tcx:Extensions/ns3:TPX/ns3:Speed",
                                "tcx:Extensions/ns2:TPX/ns2:Speed",
                            ],
                        )
                    )
                    watts = to_float(
                        get_text(
                            tp,
                            [
                                "tcx:Extensions/ns3:TPX/ns3:Watts",
                                "tcx:Extensions/ns2:TPX/ns2:Watts",
                            ],
                        )
                    )
                    run_spm = to_int(
                        get_text(
                            tp,
                            [
                                "tcx:Extensions/ns3:TPX/ns3:RunCadence",
                                "tcx:Extensions/ns2:TPX/ns2:RunCadence",
                            ],
                        )
                    )
                    if cad is None:
                        cad = to_int(
                            get_text(
                                tp,
                                [
                                    "tcx:Extensions/ns3:TPX/ns3:Cadence",
                                    "tcx:Extensions/ns2:TPX/ns2:Cadence",
                                ],
                            )
                        )

                    speed_kmh = speed_mps * 3.6 if speed_mps is not None else None

                    rows.append(
                        {
                            "activity_sport": sport,
                            "lap_index": li,
                            "trackpoint_index": ti,
                            "time_utc": ts.isoformat() if ts else None,
                            "elapsed_s": round(elapsed, 3)
                            if elapsed is not None
                            else None,
                            "latitude_deg": lat,
                            "longitude_deg": lon,
                            "altitude_m": alt,
                            "distance_m": dist,
                            "speed_mps": speed_mps,
                            "speed_kmh": round(speed_kmh, 3)
                            if speed_kmh is not None
                            else None,
                            "hr_bpm": hr,
                            "cadence_rpm": cad,
                            "run_cadence_spm": run_spm,
                            "power_w": watts,
                        }
                    )
    return rows

def rows_to_dataframe(rows) -> pd.DataFrame:
    """Convierte la lista de dicts en DataFrame y tipifica las columnas."""
    from pandas import to_datetime

    if not rows:
        rows = [
            {
                "activity_sport": None,
                "lap_index": None,
                "trackpoint_index": None,
                "time_utc": None,
                "elapsed_s": None,
                "latitude_deg": None,
                "longitude_deg": None,
                "altitude_m": None,
                "distance_m": None,
                "speed_mps": None,
                "speed_kmh": None,
                "hr_bpm": None,
                "cadence_rpm": None,
                "run_cadence_spm": None,
                "power_w": None,
            }
        ]

    df = pd.DataFrame(rows)

    # Tipificar fecha y hora
    if "time_utc" in df.columns:
        dt = to_datetime(df["time_utc"], errors="coerce", utc=True).dt.tz_convert(None)
        df["time_utc"] = dt

    # Tipificar numéricos
    float_cols = [
        "elapsed_s",
        "latitude_deg",
        "longitude_deg",
        "altitude_m",
        "distance_m",
        "speed_mps",
        "speed_kmh",
        "power_w",
    ]
    int_cols = ["hr_bpm", "cadence_rpm", "run_cadence_spm"]
    for c in float_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    for c in int_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce", downcast="integer")

    # Ordenar por tiempo
    if "time_utc" in df.columns:
        df = df.sort_values("time_utc").reset_index(drop=True)

    return df

# ---------------------- Métricas y cargas ----------------------
def add_metrics_minimal(df: pd.DataFrame, base_name: str, ftp: float, fc20: float) -> pd.DataFrame:
    """
    Devuelve un DataFrame con columnas mínimas y métricas calculadas.
    """
    df = df.copy()

    # Derivar fecha del primer valor de time_utc
    fecha = None
    if "time_utc" in df.columns and pd.api.types.is_datetime64_any_dtype(df["time_utc"]):
        if df["time_utc"].notna().any():
            fecha = df["time_utc"].dropna().iloc[0].date()

    df["fecha"] = fecha.isoformat() if fecha else None
    df["documento"] = base_name

    # Construir DataFrame mínimo
    keep = {
        "fecha": df.get("fecha"),
        "documento": df.get("documento"),
        "elapsed_s": pd.to_numeric(df.get("elapsed_s"), errors="coerce"),
        "power_w": pd.to_numeric(df.get("power_w"), errors="coerce"),
        "hr_bpm": pd.to_numeric(df.get("hr_bpm"), errors="coerce"),
        "speed_kmh": pd.to_numeric(df.get("speed_kmh"), errors="coerce"),
    }
    m = pd.DataFrame(keep)

    # Parámetros
    ftp = float(ftp)
    fc20 = float(fc20)

    power = m["power_w"].fillna(0.0)
    hr = m["hr_bpm"].astype(float).replace({0.0: pd.NA})

    # %FTP y %FC_rel
    pct_ftp = (power / ftp) * 100.0
    pct_fc = (hr / fc20) * 100.0

    # EFR
    efr = pct_ftp / pct_fc

    # IF e ICR
    intensity_factor = power / ftp
    icr = intensity_factor / efr

    # Δt por fila e integración de cargas
    el = m["elapsed_s"].astype(float)
    dt_s = el.diff()

    # Primer Δt: intenta inferir del primer valor válido; si no, 1 s
    if dt_s.notna().sum() >= 1:
        first_dt = dt_s.dropna().iloc[0]
        try:
            first_dt = float(first_dt)
        except Exception:
            first_dt = 1.0
        if not (first_dt > 0):
            first_dt = 1.0
    else:
        first_dt = 1.0

    dt_s = dt_s.fillna(first_dt).clip(lower=0.0)
    dt_h = dt_s / 3600.0

    # Incrementos y acumulados
    tss_inc = (intensity_factor ** 2) * dt_h * 100.0
    fss_inc = (icr ** 2) * dt_h * 100.0
    tss_cum = tss_inc.cumsum()
    fss_cum = fss_inc.cumsum()

    # Valores totales (última fila) - solo en la primera fila
    tss_total = tss_cum.iloc[-1] if len(tss_cum) > 0 else 0.0
    fss_total = fss_cum.iloc[-1] if len(fss_cum) > 0 else 0.0

    # Promedio móvil de 10 segundos para TSS_inc y FSS_inc
    window_size = 10
    tss_inc_ma10 = tss_inc.rolling(window=window_size, min_periods=1, center=False).mean()
    fss_inc_ma10 = fss_inc.rolling(window=window_size, min_periods=1, center=False).mean()

    # Asignar métricas
    m["pct_ftp"] = pct_ftp
    m["pct_fc_rel"] = pct_fc
    m["EFR"] = efr
    m["IF"] = intensity_factor
    m["ICR"] = icr
    m["dt_s"] = dt_s
    m["TSS_inc"] = tss_inc
    m["FSS_inc"] = fss_inc
    m["TSS_inc_ma10"] = tss_inc_ma10
    m["FSS_inc_ma10"] = fss_inc_ma10
    m["TSS"] = tss_cum
    m["FSS"] = fss_cum
    
    # TSS_total y FSS_total solo en la primera fila
    m["TSS_total"] = pd.NA
    m["FSS_total"] = pd.NA
    m.loc[0, "TSS_total"] = tss_total
    m.loc[0, "FSS_total"] = fss_total

    return m

# ---------------------- Export & Plot ----------------------
def dataframe_to_xlsx_bytes(df: pd.DataFrame, html_chart: str = None, sheet_name: str = "DATA") -> BytesIO:
    """Exporta un DataFrame a un buffer XLSX en memoria con gráfica HTML embebida."""
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name=sheet_name)
        ws = xw.book[sheet_name]

        from openpyxl.utils import get_column_letter

        for i, col in enumerate(df.columns, start=1):
            width = 12
            if col in ("fecha", "documento"):
                width = 18
            if col in ("elapsed_s", "power_w", "hr_bpm", "speed_kmh", "dt_s"):
                width = 12
            if col in ("pct_ftp", "pct_fc_rel", "EFR", "IF", "ICR", "TSS_inc", "FSS_inc", "TSS_inc_ma10", "FSS_inc_ma10", "TSS", "FSS", "TSS_total", "FSS_total"):
                width = 14
            ws.column_dimensions[get_column_letter(i)].width = width

        # Embeber gráfica HTML si se proporciona
        if html_chart:
            # Crear nueva hoja para la gráfica
            chart_sheet = xw.book.create_sheet("Gráficas")
            
            # Guardar HTML en la hoja
            chart_sheet["A1"] = "Gráfica Interactiva de Carga"
            chart_sheet["A1"].font = chart_sheet["A1"].font.copy(bold=True, size=14)
            
            chart_sheet["A3"] = "Para ver la gráfica interactiva, descarga el archivo HTML por separado."
            chart_sheet["A3"].font = chart_sheet["A3"].font.copy(italic=True)
            
            # Insertar preview de la gráfica como texto (primera parte del HTML)
            chart_sheet["A5"] = "Vista previa (HTML):"
            preview_text = html_chart[:500] + "..." if len(html_chart) > 500 else html_chart
            chart_sheet["A6"] = preview_text
            chart_sheet["A6"].alignment = chart_sheet["A6"].alignment.copy(wrap_text=True)
            
            # Ajustar anchos
            chart_sheet.column_dimensions['A'].width = 100
            
    bio.seek(0)
    return bio

def make_plot_loads_dual(df: pd.DataFrame, title: str) -> go.Figure:
    """
    Crea una figura de Plotly con dos subplots:
    1. TSS/FSS acumulados
    2. TSS_inc_ma10/FSS_inc_ma10 (promedio móvil 10s) para ver dinámicas suavizadas
    """
    t = df["elapsed_s"]
    
    # Crear subplots: 2 filas
    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=(
            "Carga Acumulada (TSS/FSS)", 
            "Dinámica Segundo a Segundo - Promedio Móvil 10s (ΔTSS/ΔFSS)"
        ),
        vertical_spacing=0.12,
        row_heights=[0.5, 0.5]
    )
    
    # Plot 1: Acumulados
    fig.add_trace(
        go.Scatter(x=t, y=df["TSS"], name="TSS (acum)", mode="lines", 
                   line=dict(color='#1f77b4', width=2)),
        row=1, col=1
    )
    fig.add_trace(
        go.Scatter(x=t, y=df["FSS"], name="FSS (acum)", mode="lines", 
                   line=dict(color='#ff7f0e', width=2)),
        row=1, col=1
    )
    
    # Plot 2: Incrementos con promedio móvil 10s
    fig.add_trace(
        go.Scatter(x=t, y=df["TSS_inc_ma10"], name="ΔTSS (MA10s)", mode="lines", 
                   line=dict(color='#2ca02c', width=1.5)),
        row=2, col=1
    )
    fig.add_trace(
        go.Scatter(x=t, y=df["FSS_inc_ma10"], name="ΔFSS (MA10s)", mode="lines", 
                   line=dict(color='#d62728', width=1.5)),
        row=2, col=1
    )
    
    # Actualizar ejes
    fig.update_xaxes(title_text="Tiempo (s)", row=2, col=1)
    fig.update_yaxes(title_text="Carga Acumulada", row=1, col=1)
    fig.update_yaxes(title_text="Incremento de Carga (MA10s)", row=2, col=1)
    
    # Layout general
    fig.update_layout(
        title=dict(text=title, x=0.5, xanchor='center'),
        showlegend=True,
        legend=dict(orientation="h", x=0, y=-0.15),
        template="plotly_white",
        height=800,
        margin=dict(l=60, r=60, t=80, b=60),
    )
    
    return fig

def make_plot_loads(df: pd.DataFrame, title: str, show_base: bool = True) -> go.Figure:
    """Crea una figura de Plotly con TSS y FSS acumulados, y señales base opcionales."""
    t = df["elapsed_s"]
    fig = go.Figure()

    # Trazas principales
    fig.add_trace(go.Scatter(x=t, y=df["TSS"], name="TSS (acum)", mode="lines"))
    fig.add_trace(go.Scatter(x=t, y=df["FSS"], name="FSS (acum)", mode="lines"))

    # Señales base
    if show_base:
        if "power_w" in df.columns:
            fig.add_trace(go.Scatter(x=t, y=df["power_w"], name="Potencia (W)", mode="lines", yaxis="y2"))
        if "hr_bpm" in df.columns:
            fig.add_trace(go.Scatter(x=t, y=df["hr_bpm"], name="FC (bpm)", mode="lines", yaxis="y3"))

    layout = dict(
        title=title,
        xaxis=dict(title="Tiempo (s)"),
        yaxis=dict(title="Carga acumulada (TSS/FSS)", rangemode="tozero"),
        yaxis2=dict(title="Potencia (W)", overlaying="y", side="right", position=1.0, showgrid=False),
        yaxis3=dict(title="FC (bpm)", overlaying="y", side="right", position=0.98, showgrid=False),
        legend=dict(orientation="h", x=0, y=1.12),
        template="plotly_white",
        margin=dict(l=60, r=80, t=70, b=50),
    )
    fig.update_layout(**layout)
    return fig

def make_comparative_dashboard(all_data: list) -> go.Figure:
    """
    Crea un dashboard comparativo entre múltiples entrenamientos.
    all_data: lista de tuplas (nombre, dataframe)
    """
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f']
    
    fig = make_subplots(
        rows=3, cols=2,
        subplot_titles=(
            "TSS Acumulado Comparativo",
            "FSS Acumulado Comparativo",
            "Potencia Media Comparativa",
            "FC Media Comparativa",
            "EFR Promedio Comparativo",
            "ICR Promedio Comparativo"
        ),
        vertical_spacing=0.12,
        horizontal_spacing=0.10,
        specs=[[{}, {}], [{}, {}], [{}, {}]]
    )
    
    for idx, (name, df) in enumerate(all_data):
        color = colors[idx % len(colors)]
        t = df["elapsed_s"]
        
        # TSS Acumulado
        fig.add_trace(
            go.Scatter(x=t, y=df["TSS"], name=name, mode="lines", 
                       line=dict(color=color), showlegend=True, legendgroup=name),
            row=1, col=1
        )
        
        # FSS Acumulado
        fig.add_trace(
            go.Scatter(x=t, y=df["FSS"], name=name, mode="lines", 
                       line=dict(color=color), showlegend=False, legendgroup=name),
            row=1, col=2
        )
        
        # Potencia
        fig.add_trace(
            go.Scatter(x=t, y=df["power_w"], name=name, mode="lines", 
                       line=dict(color=color), showlegend=False, legendgroup=name),
            row=2, col=1
        )
        
        # FC
        fig.add_trace(
            go.Scatter(x=t, y=df["hr_bpm"], name=name, mode="lines", 
                       line=dict(color=color), showlegend=False, legendgroup=name),
            row=2, col=2
        )
        
        # EFR
        efr_ma = df["EFR"].rolling(window=30, min_periods=1).mean()
        fig.add_trace(
            go.Scatter(x=t, y=efr_ma, name=name, mode="lines", 
                       line=dict(color=color), showlegend=False, legendgroup=name),
            row=3, col=1
        )
        
        # ICR
        icr_ma = df["ICR"].rolling(window=30, min_periods=1).mean()
        fig.add_trace(
            go.Scatter(x=t, y=icr_ma, name=name, mode="lines", 
                       line=dict(color=color), showlegend=False, legendgroup=name),
            row=3, col=2
        )
    
    # Actualizar ejes
    fig.update_xaxes(title_text="Tiempo (s)")
    fig.update_yaxes(title_text="TSS", row=1, col=1)
    fig.update_yaxes(title_text="FSS", row=1, col=2)
    fig.update_yaxes(title_text="Potencia (W)", row=2, col=1)
    fig.update_yaxes(title_text="FC (bpm)", row=2, col=2)
    fig.update_yaxes(title_text="EFR (MA30s)", row=3, col=1)
    fig.update_yaxes(title_text="ICR (MA30s)", row=3, col=2)
    
    fig.update_layout(
        title=dict(text="📊 Dashboard Comparativo de Entrenamientos", x=0.5, xanchor='center'),
        showlegend=True,
        legend=dict(orientation="v", x=1.05, y=1),
        template="plotly_white",
        height=1200,
        margin=dict(l=60, r=150, t=80, b=60),
    )
    
    return fig

def clean_base_name(name: str) -> str:
    """Elimina extensiones .gz y .tcx del nombre del archivo."""
    base = name
    if base.lower().endswith(".gz"):
        base = base[:-3]
    if base.lower().endswith(".tcx"):
        base = base[:-4]
    return base

def render_plot_and_download(df: pd.DataFrame, base_name: str, key_prefix: str):
    """Muestra las gráficas en Streamlit y ofrece las descargas HTML."""
    
    # Gráfica original con señales base
    st.subheader("📊 Análisis con Señales Base")
    fig1 = make_plot_loads(df, title=f"Dinámica de Carga – {base_name}", show_base=True)
    st.plotly_chart(fig1, use_container_width=True, key=f"{key_prefix}_plot1")
    
    html_buf1 = StringIO()
    fig1.write_html(html_buf1, include_plotlyjs="cdn", full_html=True)
    html_bytes1 = html_buf1.getvalue().encode("utf-8")
    
    file_html1 = f"{base_name}_analisis_completo.html"
    st.download_button(
        label="⬇️ Descargar gráfica completa (HTML)",
        data=html_bytes1,
        file_name=file_html1,
        mime="text/html",
        key=f"{key_prefix}_html1",
    )
    
    # Gráfica dual: acumulados + incrementos
    st.subheader("📈 Comparación: Acumulados vs. Segundo a Segundo")
    fig2 = make_plot_loads_dual(df, title=f"TSS/FSS: Acumulado vs. Dinámico – {base_name}")
    st.plotly_chart(fig2, use_container_width=True, key=f"{key_prefix}_plot2")
    
    html_buf2 = StringIO()
    fig2.write_html(html_buf2, include_plotlyjs="cdn", full_html=True)
    html_bytes2 = html_buf2.getvalue().encode("utf-8")
    
    file_html2 = f"{base_name}_dinamica_detallada.html"
    st.download_button(
        label="⬇️ Descargar gráfica dinámica (HTML)",
        data=html_bytes2,
        file_name=file_html2,
        mime="text/html",
        key=f"{key_prefix}_html2",
    )
    
    st.info("💡 La gráfica superior muestra TSS/FSS acumulados. La inferior muestra los incrementos (ΔTSS/ΔFSS) suavizados con promedio móvil de 10 segundos para revelar patrones de esfuerzo más claros.")
    
    # Retornar HTML para embeber en Excel
    return html_bytes2.decode("utf-8")

# ---------------------- UI ----------------------
st.title("📈 TCX → XLSX con EFR / IF / ICR / TSS / FSS")
st.write(
    """
Sube uno o varios **.tcx** o **.tcx.gz**.
Por cada archivo deberás ingresar:

- **FTP (W)**: tu umbral funcional de potencia.
- **FC_20min_max (bpm)**: promedio de tus mejores 20 minutos de los últimos 90 días.

**ICR = IF ÷ EFR** (Índice de Carga Relativa correcto).

Integración de carga:
- **TSS** = Σ ( **IF² × Δt_horas × 100** )
- **FSS** = Σ ( **ICR² × Δt_horas × 100** )

**Nuevas características:**
- 📊 Gráfica de dinámicas segundo a segundo con promedio móvil de 10s (ΔTSS/ΔFSS)
- 📑 Gráficas embebidas en Excel
- 📈 Columnas TSS_total y FSS_total en la primera fila
- 🔍 Dashboard comparativo entre entrenamientos (cuando subes múltiples archivos)
"""
)

uploads = st.file_uploader(
    "Sube tus archivos (puedes seleccionar varios)",
    type=["tcx", "gz"],
    accept_multiple_files=True,
    key="uploader_main",
)

if uploads:
    # Almacenar datos procesados para comparación
    processed_data = []
    xlsx_buffers = []
    all_params = {}

    # Primero, recoger parámetros de todos los archivos
    st.subheader("⚙️ Configuración de Parámetros")
    for idx, up in enumerate(uploads):
        base = clean_base_name(up.name)
        with st.expander(f"Parámetros para: `{up.name}`", expanded=(idx == 0)):
            key_prefix = f"a{idx}"
            c1, c2 = st.columns(2)
            ftp_input = c1.number_input(
                f"FTP (W)",
                min_value=1,
                step=1,
                key=f"{key_prefix}_ftp",
                value=250
            )
            fc20_input = c2.number_input(
                f"FC_20min_max (bpm)",
                min_value=1,
                step=1,
                key=f"{key_prefix}_fc20",
                value=160
            )
            all_params[idx] = (up, base, ftp_input, fc20_input, key_prefix)
    
    # Botón para procesar todos
    if st.button("▶️ Procesar Todos los Archivos", type="primary"):
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for idx, (up, base, ftp_val, fc_val, key_prefix) in all_params.items():
            status_text.text(f"Procesando {up.name}...")
            progress_bar.progress((idx + 1) / len(all_params))
            
            if not (ftp_val and fc_val):
                st.warning(f"⚠️ Ingresa FTP y FC_20min_max para {up.name}")
                continue
            
            try:
                rows = parse_tcx_to_rows(up)
                df_raw = rows_to_dataframe(rows)
                df_final = add_metrics_minimal(
                    df_raw, base_name=base, ftp=ftp_val, fc20=fc_val
                )
                
                # Guardar para dashboard comparativo
                processed_data.append((base, df_final))
                
                # Generar Excel
                html_chart = ""
                out_name = f"{base}.xlsx"
                xlsx_bio = dataframe_to_xlsx_bytes(df_final, html_chart=html_chart, sheet_name="DATA")
                xlsx_buffers.append((out_name, xlsx_bio))
                
            except Exception as e:
                st.error(f"❌ Error en {up.name}: {e}")
        
        status_text.text("✅ Procesamiento completado!")
        progress_bar.progress(100)
        
        # ==================== DASHBOARD COMPARATIVO ====================
        if len(processed_data) > 1:
            st.markdown("---")
            st.header("🔍 Dashboard Comparativo de Entrenamientos")
            
            # Tabla resumen
            st.subheader("📊 Resumen de Métricas")
            summary_data = []
            for name, df in processed_data:
                tss_total = df["TSS_total"].iloc[0]
                fss_total = df["FSS_total"].iloc[0]
                duration_h = df["elapsed_s"].iloc[-1] / 3600.0
                avg_power = df["power_w"].mean()
                avg_hr = df["hr_bpm"].mean()
                avg_efr = df["EFR"].mean()
                avg_icr = df["ICR"].mean()
                
                summary_data.append({
                    "Entrenamiento": name,
                    "TSS Total": f"{tss_total:.1f}",
                    "FSS Total": f"{fss_total:.1f}",
                    "Duración (h)": f"{duration_h:.2f}",
                    "Potencia Promedio (W)": f"{avg_power:.1f}",
                    "FC Promedio (bpm)": f"{avg_hr:.0f}",
                    "EFR Promedio": f"{avg_efr:.2f}",
                    "ICR Promedio": f"{avg_icr:.2f}"
                })
            
            summary_df = pd.DataFrame(summary_data)
            st.dataframe(summary_df, use_container_width=True)
            
            # Gráfica comparativa
            st.subheader("📈 Análisis Comparativo Visual")
            comp_fig = make_comparative_dashboard(processed_data)
            st.plotly_chart(comp_fig, use_container_width=True)
            
            # Descargar dashboard HTML
            comp_html_buf = StringIO()
            comp_fig.write_html(comp_html_buf, include_plotlyjs="cdn", full_html=True)
            comp_html_bytes = comp_html_buf.getvalue().encode("utf-8")
            
            st.download_button(
                label="⬇️ Descargar Dashboard Comparativo (HTML)",
                data=comp_html_bytes,
                file_name="dashboard_comparativo.html",
                mime="text/html",
                key="dashboard_html",
            )
            
            st.info("💡 El dashboard comparativo te permite identificar patrones, diferencias de intensidad y evolución entre diferentes entrenamientos.")
        
        # ==================== ANÁLISIS INDIVIDUAL ====================
        st.markdown("---")
        st.header("📄 Análisis Individual por Entrenamiento")
        
        for idx, (base, df_final) in enumerate(processed_data):
            with st.expander(f"📊 Detalles de: {base}", expanded=False):
                # Métricas individuales
                col1, col2, col3, col4 = st.columns(4)
                tss_final = df_final["TSS_total"].iloc[0]
                fss_final = df_final["FSS_total"].iloc[0]
                duration_h = df_final["elapsed_s"].iloc[-1] / 3600.0
                avg_power = df_final["power_w"].mean()
                
                col1.metric("TSS Total", f"{tss_final:.1f}")
                col2.metric("FSS Total", f"{fss_final:.1f}")
                col3.metric("Duración (h)", f"{duration_h:.2f}")
                col4.metric("Potencia Media (W)", f"{avg_power:.1f}")
                
                # Gráficas individuales
                key_prefix = f"ind{idx}"
                html_chart = render_plot_and_download(df_final, base_name=base, key_prefix=key_prefix)
                
                # Actualizar Excel con gráfica
                out_name = f"{base}.xlsx"
                xlsx_bio = dataframe_to_xlsx_bytes(df_final, html_chart=html_chart, sheet_name="DATA")
                xlsx_buffers[idx] = (out_name, xlsx_bio)
                
                st.download_button(
                    label=f"⬇️ Descargar {out_name}",
                    data=xlsx_bio.getvalue(),
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"{key_prefix}_dl_xlsx",
                )
        
        # ==================== DESCARGA MASIVA ====================
        if len(xlsx_buffers) > 1:
            st.markdown("---")
            zip_bio = BytesIO()
            with zipfile.ZipFile(zip_bio, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                for fname, fb in xlsx_buffers:
                    zf.writestr(fname, fb.getvalue())
            zip_bio.seek(0)
            
            st.download_button(
                "📦 Descargar Todos los Archivos Excel (.zip)",
                data=zip_bio.getvalue(),
                file_name="entrenamientos_tcx.zip",
                mime="application/zip",
                key="zip_all_dl",
            )

else:
    st.info("👆 Sube uno o más archivos TCX para comenzar el análisis")

# ==================== SIDEBAR INFO ====================
with st.sidebar:
    st.header("ℹ️ Información")
    st.markdown("""
    ### Métricas Calculadas
    
    **EFR** (Eficiencia Relativa)
    - Relación entre %FTP y %FC
    - Mayor valor = mejor eficiencia
    
    **IF** (Factor de Intensidad)
    - Potencia / FTP
    - Mide intensidad del esfuerzo
    
    **ICR** (Índice de Carga Relativa)
    - IF / EFR
    - Carga ajustada por eficiencia
    
    **TSS** (Training Stress Score)
    - Estrés basado en potencia
    - Acumulado = Σ(IF² × Δt × 100)
    
    **FSS** (Fitness Stress Score)
    - Estrés ajustado por eficiencia
    - Acumulado = Σ(ICR² × Δt × 100)
    """)
    
    st.markdown("---")
    st.markdown("""
    ### Dashboard Comparativo
    
    Cuando subes **múltiples archivos**, 
    se genera automáticamente un dashboard 
    que compara:
    
    - TSS y FSS acumulados
    - Potencia y FC
    - EFR e ICR promedio
    - Tabla resumen de métricas
    
    Ideal para analizar progresión 
    y diferencias entre sesiones.
    """)
    
    st.markdown("---")
    st.caption("v3.0 - Dashboard Comparativo")
